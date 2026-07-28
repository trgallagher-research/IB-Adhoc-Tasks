#!/usr/bin/env python3
"""Regenerate normalized source inventories from the sanitized evidence files.

Inputs (evidence, read-only):
  01-forms-excel/sanitized/Innovation-Intake-Form-responses-reference.xlsx
  02-get-response-details/sanitized/get-response-details-response-6.body.json

Outputs (regenerated, do not hand-edit):
  01-forms-excel/sanitized/forms-question-inventory.json / .md
  02-get-response-details/sanitized/response-keys-inventory.json / .md

Policy: structural inventory only. Answer values are dummy test data. The
Email/Name metadata column *values* are never reproduced.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "01-forms-excel/sanitized/Innovation-Intake-Form-responses-reference.xlsx"
BODY = ROOT / "02-get-response-details/sanitized/get-response-details-response-6.body.json"
GENERATED = "2026-07-28"

METADATA = {1: "ID", 2: "Start time", 3: "Completion time", 4: "Email",
            5: "Name", 6: "Last modified time"}

COLUMN_NOTES = {
    9: ["date question; Excel shows datetime at midnight"],
    12: ["conditional on 'External Partner Involved?' = Yes (filled only in response 2)"],
    13: ["conditional on 'External Partner Involved?' = Yes (filled only in response 2)"],
    14: ["conditional on 'External Partner Involved?' = Yes (filled only in response 2)"],
    15: ["multi-choice; Excel serializes as 'Choice1;Choice2;' with trailing semicolon"],
    18: ["conditional on 'Does this suggested idea directly impact a local market?' = Yes"],
    20: ["conditional on 'Is a compliance boundary adaptation required?' = Yes"],
    21: ["conditional on 'Is a compliance boundary adaptation required?' = Yes"],
    22: ["BLANK in all 6 reference responses including fully-completed response 2 — "
         "almost certainly a display-only notice element with no input"],
    23: ["filled only in responses 1-2 (both local market = Yes); appears conditional "
         "on the local-market branch — unproven, verify against live form"],
    24: ["filled only in responses 1-2; see column 23 note"],
    25: ["filled only in responses 1-2; see column 23 note"],
    26: ["filled only in responses 1-2; see column 23 note"],
    30: ["multi-choice; Excel serializes as 'Choice1;Choice2;' with trailing semicolon"],
    42: ["conditional on IBEN impact = Yes"],
    44: ["conditional on Professional Learning impact = Yes"],
    46: ["conditional on additional factors = Yes"],
    47: ["file-upload question; blank in all reference responses"],
}


def norm(s):
    return s.replace("\xa0", " ").strip() if isinstance(s, str) else s


def build_forms_inventory():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    rows = [[c.value for c in ws[r]] for r in range(2, ws.max_row + 1)]

    cols = []
    for i, h in enumerate(hdr, 1):
        vals = [row[i - 1] for row in rows if row[i - 1] not in (None, "")]
        notes = list(COLUMN_NOTES.get(i, []))
        if vals and all(str(v) in "12345" for v in vals) and i not in METADATA:
            notes.append("Likert-style rating; observed integer strings within 1-5")
        cols.append({
            "excel_column": i,
            "label": norm(h),
            "label_raw": h,
            "kind": "forms_metadata" if i in METADATA else "question",
            "responses_with_value": len(vals),
            "observed_python_types": sorted({type(v).__name__ for v in vals}),
            "notes": notes,
        })

    inv = {
        "_provenance": {
            "source": str(XLSX.relative_to(ROOT)),
            "generated_by": "scripts/build_inventories.py",
            "generated": GENERATED,
            "content_policy": ("Structural inventory only. Answer values are dummy test "
                               "data; no real submission content. Personal metadata values "
                               "(Email/Name columns) are NOT reproduced here."),
            "totals": {"columns": len(hdr),
                       "forms_metadata_columns": len(METADATA),
                       "question_columns": len(hdr) - len(METADATA),
                       "reference_responses": len(rows)},
        },
        "columns": cols,
    }
    out = ROOT / "01-forms-excel/sanitized/forms-question-inventory.json"
    out.write_text(json.dumps(inv, indent=2, ensure_ascii=False) + "\n")

    md = ["# Forms question inventory (normalized)", "",
          f"Generated {GENERATED} by `scripts/build_inventories.py` from the dummy-safe "
          "Excel reference export. 47 columns = 6 Forms metadata + 41 question/output "
          "columns, across 6 dummy reference responses.", "",
          "| # | Kind | Question label | Filled (of 6) | Notes |",
          "|---|------|----------------|---------------|-------|"]
    for c in cols:
        md.append(f"| {c['excel_column']} | {c['kind'].replace('forms_', '')} | "
                  f"{c['label'][:90]} | {c['responses_with_value']} | "
                  f"{'; '.join(c['notes'])} |")
    (ROOT / "01-forms-excel/sanitized/forms-question-inventory.md").write_text(
        "\n".join(md) + "\n")
    return inv


def build_key_inventory():
    body = json.loads(BODY.read_text())["body"]
    rkeys = [k for k in body if re.fullmatch(r"r[0-9a-f]{32}", k)]
    keys = []
    for k in rkeys:
        v = body[k]
        if v == "":
            shape = "blank"
        elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
            shape = "iso-date"
        elif v in ("Yes", "No"):
            shape = "yes-no"
        elif v.isdigit():
            shape = "numeric-string"
        elif v.startswith("["):
            shape = "json-array-string (multi-choice serialization)"
        else:
            shape = "free-text"
        keys.append({"response_key": k,
                     "observed_value_response_6": v,
                     "value_shape": shape})

    kinv = {
        "_provenance": {
            "source": str(BODY.relative_to(ROOT)),
            "generated_by": "scripts/build_inventories.py",
            "generated": GENERATED,
            "content_policy": ("All values are dummy test content from Form response ID 6. "
                               "Keys are structural identifiers retained deliberately."),
            "totals": {"opaque_r_keys": len(rkeys),
                       "non_blank": sum(1 for k in keys if k["value_shape"] != "blank"),
                       "blank": sum(1 for k in keys if k["value_shape"] == "blank"),
                       "non_key_properties": ["responder", "submitDate"]},
        },
        "keys": keys,
    }
    out = ROOT / "02-get-response-details/sanitized/response-keys-inventory.json"
    out.write_text(json.dumps(kinv, indent=2, ensure_ascii=False) + "\n")

    md = ["# `Get response details` response-key inventory", "",
          f"Generated {GENERATED} by `scripts/build_inventories.py` from the sanitized "
          "response-6 body. 48 opaque `r…` keys plus `responder` and `submitDate`. "
          "All values shown are dummy test content.", "",
          "| Response key | Shape | Observed value (dummy, response 6) |",
          "|--------------|-------|------------------------------------|"]
    for k in keys:
        v = k["observed_value_response_6"]
        md.append(f"| `{k['response_key']}` | {k['value_shape']} | "
                  f"{v[:80] if v else '—'} |")
    (ROOT / "02-get-response-details/sanitized/response-keys-inventory.md").write_text(
        "\n".join(md) + "\n")
    return kinv


if __name__ == "__main__":
    inv = build_forms_inventory()
    kinv = build_key_inventory()
    print("forms columns:", inv["_provenance"]["totals"])
    print("response keys:", kinv["_provenance"]["totals"])
